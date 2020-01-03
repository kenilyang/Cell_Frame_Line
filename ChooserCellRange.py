def ccr(start_cell,end_cell,merg,wsdata,ifl_style,ofl_style,ftcolor,bgcolor):
	import openpyxl
	from openpyxl import workbook, load_workbook
	from openpyxl import Workbook
	from openpyxl.styles import Font, colors, Alignment, PatternFill, Border, Side
	import os
	from os import listdir
	from os.path import isfile, isdir, join
	import time,datetime
	lm=[]
	if isfile('abc.xlsx'):					# 如果 'abc.xlsx' 這個檔案存在
		wb = load_workbook('abc.xlsx')	# 就開啟該檔
	else:							# 否則
		wb = Workbook()					# 以寫入模式開啟
	ws = wb.active						# 將做用中的 Excel 指定到 ws 變數
	ws.title = 'sh001'					# 指定表單名稱
	wb.save('abc.xlsx')
	lista = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ' 			# 為了方便以下程式的執行，此清單可用來指定儲存格的名稱

	# 以下為框線、字型、文字位置的預設值
	font = Font(name='Times New Roman', size=12, italic=False, color='000000', bold=False)		# 設定字型
	alignment = Alignment(horizontal='center', vertical='center', wrapText=True)			# 儲存格內容定位，預設置中
	border = Border(left=Side(border_style='thin',color='000000'),\
					right=Side(border_style='thin',color='000000'),\
					top=Side(border_style='thin',color='000000'),\
					bottom=Side(border_style='thin',color='000000'))							# 設置 "預設表框" 為細框線

	#  以下判斷儲存格名稱是否正確 
	if not start_cell[0].isalpha() or not start_cell[-1].isdigit():			# 如果 start_cell 的第一個字元不是英文字母，或，最後一個字元不是數字
		os.system('cls')
		print('start_cell = ' + start_cell + '，參數輸入錯誤，請確認.....')	# 就顯示錯誤
		os.system('pause')
		exit()																# 並離開程式
	elif not end_cell[0].isalpha() or not end_cell[-1].isdigit():			# 如果  end_cell  的第一個字元不是英文字母，或，最後一個字元不是數字
		os.system('cls')
		print('end_cell = ' + end_cell + '，參數輸入錯誤，請確認.....')		# 就顯示錯誤
		os.system('pause')
		exit()																# 並離開程式
	else :
		sa = start_cell[0].upper()			# 將 start_cell 第一個字元，存入 sa，並強制 start_cell 的第一個字元為大寫
		sn = int(start_cell[1:len(start_cell) + 1])	# 將 start_cell 第二個字元 到 最後一個字元存入 sn ，並強制為數字
		ea = end_cell[0].upper()			# 將  end_cell  第一個字元，存入 ea，並強制  end_cell  的第一個字元為大寫
		en = int(end_cell[1:len(end_cell) + 1])		# 將  end_cell  第二個字元 到 最後一個字元存入 en ，並強制為數字


	#  以下判斷是否合併，並將 wsdata 內容存入第一個儲存格(也就是start_cell)
	if merg == 'y' or merg == 'Y':
		ws.merge_cells(str(sa) + str(sn) + ':' + str(ea) + str(en))	# 合併指令，範例：ws.merge_cells('A1:D6')
		#ws[str(sa) + str(sn)] = wsdata
		ws[start_cell] = wsdata
	sa = lista.index(sa) + 1						# 尋找 sa 字元在 lista 裡的序號是多少，並 +1，因為 range 不能為 0
	ea = lista.index(ea) + 1						# 尋找 ea 字元在 lista 裡的序號是多少，並 +1，因為 range 不能為 0
	
	# 以下為選擇儲存格範圍的主程式，因為有行、列兩種選擇，所以用兩個迴圈來做
	for l in range(sa,ea + 1): 				# 設定 小寫的 L 為行，即儲存格 "英文字母" 的部份
		for m in range(sn,en + 1): 			# 設定 m 為列，即儲存格 "數字" 的部份
			lm.append(str(lista[l-1]) + str(m))	# 將範圍內的所有 "儲存格名稱" 存成清單 lm

			# 以下是設定儲存格的預設格式
			ws.cell(row=m, column=l).alignment = alignment		# 上下左右皆置中
			ws.cell(row=m, column=l).border = border		# 內、外框線預設為細
			ws.cell(row=m, column=l).fill = PatternFill("solid", fgColor=bgcolor)	#背景顏色
			ws.cell(row=m, column=l).font = Font(name='Times New Roman', size=12, italic=False, color=ftcolor, bold=False)	#設定字型

			#此段為單一儲存格(點)
			if sa == ea and sn == en :	
				ws.cell(row=m, column=l, value= wsdata).border = Border(left=Side(border_style=ofl_style,color='000000'),\
																		right=Side(border_style=ofl_style,color='000000'),\
																		top=Side(border_style=ofl_style,color='000000'),\
																		bottom=Side(border_style=ofl_style,color='000000'))				
			#此段為整行範圍(線)
			elif sa == ea and sn != en:	
				if m == sn:	
					ws.cell(row=m, column=l).border = Border(left=Side(border_style=ofl_style,color='000000'),\
															right=Side(border_style=ofl_style,color='000000'),\
															top=Side(border_style=ofl_style,color='000000'),\
															bottom=Side(border_style=ifl_style,color='000000'))
				elif m != sn and m != en:
					ws.cell(row=m, column=l).border = Border(left=Side(border_style=ofl_style,color='000000'),\
															right=Side(border_style=ofl_style,color='000000'),\
															top=Side(border_style=ifl_style,color='000000'),\
															bottom=Side(border_style=ifl_style,color='000000'))
				else:
					ws.cell(row=m, column=l).border = Border(left=Side(border_style=ofl_style,color='000000'),\
															right=Side(border_style=ofl_style,color='000000'),\
															top=Side(border_style=ifl_style,color='000000'),\
															bottom=Side(border_style=ofl_style,color='000000'))
			
			#此段為整列範圍(線)
			elif sa != ea and sn == en:	
				if l == sa:
					ws.cell(row=m, column=l).border = Border(left=Side(border_style=ofl_style,color='000000'),\
															right=Side(border_style=ifl_style,color='000000'),\
															top=Side(border_style=ofl_style,color='000000'),\
															bottom=Side(border_style=ofl_style,color='000000'))
				elif l == ea:
					ws.cell(row=m, column=l).border = Border(left=Side(border_style=ifl_style,color='000000'),\
															right=Side(border_style=ofl_style,color='000000'),\
															top=Side(border_style=ofl_style,color='000000'),\
															bottom=Side(border_style=ofl_style,color='000000'))
				else:
					ws.cell(row=m, column=l).border = Border(left=Side(border_style=ifl_style,color='000000'),\
															right=Side(border_style=ifl_style,color='000000'),\
															top=Side(border_style=ofl_style,color='000000'),\
															bottom=Side(border_style=ofl_style,color='000000'))
				
			#此段為大範圍(面)
			else:
				if l == sa and m == sn:
					ws.cell(row=m, column=l).border = Border(left=Side(border_style=ofl_style,color='000000'),\
															right=Side(border_style=ifl_style,color='000000'),\
															top=Side(border_style=ofl_style,color='000000'),\
															bottom=Side(border_style=ifl_style,color='000000'))
				elif l == ea and m == sn:
					ws.cell(row=m, column=l).border = Border(left=Side(border_style=ifl_style,color='000000'),\
															right=Side(border_style=ofl_style,color='000000'),\
															top=Side(border_style=ofl_style,color='000000'),\
															bottom=Side(border_style=ifl_style,color='000000'))
				elif l == sa and m == en:
					ws.cell(row=m, column=l).border = Border(left=Side(border_style=ofl_style,color='000000'),\
															right=Side(border_style=ifl_style,color='000000'),\
															top=Side(border_style=ifl_style,color='000000'),\
															bottom=Side(border_style=ofl_style,color='000000'))
				elif l == ea  and m == en:
					ws.cell(row=m, column=l).border = Border(left=Side(border_style=ifl_style,color='000000'),\
															right=Side(border_style=ofl_style,color='000000'),\
															top=Side(border_style=ifl_style,color='000000'),\
															bottom=Side(border_style=ofl_style,color='000000'))
				elif l == sa and m != sn:
					ws.cell(row=m, column=l).border = Border(left=Side(border_style=ofl_style,color='000000'),\
															right=Side(border_style=ifl_style,color='000000'),\
															top=Side(border_style=ifl_style,color='000000'),\
															bottom=Side(border_style=ifl_style,color='000000'))
				elif l == ea and m != sn:
					ws.cell(row=m, column=l).border = Border(left=Side(border_style=ifl_style,color='000000'),\
															right=Side(border_style=ofl_style,color='000000'),\
															top=Side(border_style=ifl_style,color='000000'),\
															bottom=Side(border_style=ifl_style,color='000000'))
				elif l != sa and m == sn:
					ws.cell(row=m, column=l).border = Border(left=Side(border_style=ifl_style,color='000000'),\
															right=Side(border_style=ifl_style,color='000000'),\
															top=Side(border_style=ofl_style,color='000000'),\
															bottom=Side(border_style=ifl_style,color='000000'))
				elif l != ea and m == en:
					ws.cell(row=m, column=l).border = Border(left=Side(border_style=ifl_style,color='000000'),\
															right=Side(border_style=ifl_style,color='000000'),\
															top=Side(border_style=ifl_style,color='000000'),\
															bottom=Side(border_style=ofl_style,color='000000'))
	wb.save('abc.xlsx')
	return lm 	# 將 lm 清單 (範圍內的所有儲存格名稱)，回傳給 Function
